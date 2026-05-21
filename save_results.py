# save_results.py — writer de resultados (HTML/TXT) com layout centrado, dark-mode e highlight robusto
from __future__ import annotations

import os
import re
import html
import platform
import unicodedata
from datetime import datetime
from collections import defaultdict
from typing import List, Any, Iterable, Tuple, Optional

# =========================
# Utils e fallbacks
# =========================

def _try_import_boolean_parser():
    try:
        from boolean_parser import BooleanSearchParser  # type: ignore
        return BooleanSearchParser
    except Exception:
        return None

BooleanSearchParser = _try_import_boolean_parser()

def _as_attr(obj: Any, name: str, default=None):
    """Acede a atributo/chave indiferentemente."""
    if isinstance(obj, dict):
        return obj.get(name, default)
    return getattr(obj, name, default)

def _escape_path_for_js(abs_path: str) -> str:
    """Escapar para JS (Windows com backslashes)."""
    return abs_path.replace("\\", "\\\\") if platform.system() == "Windows" else abs_path

def _position_label(filepath: str, location: Any) -> Tuple[str, str]:
    """(label, page_param) para o relatório."""
    fp = (filepath or "").lower()
    if fp.endswith(".pdf"):   return (f"Page {location}", str(location) if location else "")
    if fp.endswith(".docx"):  return (f"Paragraph {location}", "")
    if fp.endswith((".html", ".htm")): return (f"Element {location}", "")
    if fp.endswith(".md"):    return (f"Block {location}", "")
    if fp.endswith((".xlsx", ".xls")):
        # Location format: "SheetName:Row:Column" or just row/col
        if isinstance(location, str) and ":" in location:
            return (f"Cell {location}", "")
        return (f"Cell {location}", "")
    if fp.endswith(".csv"):
        # Location format: "R{row}:C{col}" or "R{row}"
        if isinstance(location, str):
            return (f"Location {location}", "")
        return (f"Row {location}", "")
    return (f"Line {location}", "")

def _unique_key_for_dedup(r: Any) -> Tuple[str, str]:
    """Chave de deduplicação por (contexto, query)."""
    return (_as_attr(r, "context", "") or "", _as_attr(r, "query", "") or "")

# ---------- Normalização “de exibição” (não remove acentos) ----------
def _display_norm(s: str) -> str:
    """
    Normalização para mostrar no HTML/TXT sem perder acentos.
    Mantém \n, remove zero-width, junta palavras partidas em fim de linha.
    """
    if not isinstance(s, str):
        s = str(s or "")
    # NFKC e remoção de zero-width (mas preserva acentos)
    ZW = dict.fromkeys(map(ord, "\u200B\u200C\u200D\u200E\u200F\u2060"), None)
    s = unicodedata.normalize("NFKC", s).translate(ZW)
    # juntar palavras quebradas por newline com hífen suave ou '-'
    s = re.sub(r"(\w)[\u00AD-]\s*\n\s*(\w)", r"\1\2", s)
    # normalizar quebras
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # comprimir espaços horizontais mas manter \n
    s = re.sub(r"[ \t]+", " ", s)
    # limitar múltiplas linhas vazias
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

# ---------- Highlight robusto (frases + wildcards), case- e accent-insensitive ----------
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def _token_spans_for_accentless_map(src: str) -> Tuple[str, List[int]]:
    """
    Devolve (accentless_string, map_idx), onde map_idx[i] = índice no original.
    Útil para casar no texto sem acentos e depois mapear para os índices reais.
    """
    acc_less = []
    map_idx: List[int] = []
    for i, ch in enumerate(src):
        base = ''.join(c for c in unicodedata.normalize("NFD", ch) if unicodedata.category(c) != "Mn")
        if base:
            # normalmente 1-para-1; se base tiver >1 chars (raro), repetimos o índice
            for _ in base:
                acc_less.append(_)
                map_idx.append(i)
    return "".join(acc_less), map_idx

def _build_term_patterns(terms: Iterable[str]) -> List[re.Pattern]:
    """
    Constrói padrões regex sobre texto *sem acentos*:
      - frases entre aspas aceitam espaços flexíveis (\s+)
      - wildcards: * -> \w*, ? -> \w
      - termos simples com limites de palavra (\b)
    """
    pats: List[re.Pattern] = []
    for t in (terms or []):
        if not t:
            continue
        tt = t.strip()
        is_phrase = (tt.startswith('"') and tt.endswith('"')) or (tt.startswith("'") and tt.endswith("'"))
        core = tt[1:-1] if is_phrase else tt
        core = _strip_accents(core)
        core = re.escape(core)
        core = core.replace(r"\*", r"\w*").replace(r"\?", r"\w")
        if is_phrase:
            core = core.replace(r"\ ", r"\s+")
            rx = re.compile(core, flags=re.IGNORECASE | re.UNICODE)
        else:
            rx = re.compile(r"\b" + core + r"\b", flags=re.IGNORECASE | re.UNICODE)
        pats.append(rx)
    return pats

def _collect_terms_for_highlight(query: str) -> List[str]:
    """Extrai termos/frases (ordem preservada) da query para highlight."""
    if not query:
        return []
    if BooleanSearchParser is not None:
        try:
            p = BooleanSearchParser(query)
            ts = list(getattr(p, "search_terms", []) or [])
            # únicos case-insensitive
            seen = set(); out = []
            for t in ts:
                tl = (t or "").lower()
                if tl and tl not in seen:
                    seen.add(tl); out.append(t)
            return out
        except Exception:
            pass
    # fallback simples: remove operadores, respeita frases
    q = re.sub(r'\b(AND|OR|NOT)\b', ' ', query, flags=re.I)
    q = re.sub(r'\bNEAR/\d+\b', ' ', q, flags=re.I)
    q = q.replace("(", " ").replace(")", " ")
    parts: List[str] = []
    buf: List[str] = []; inq = False
    for ch in q:
        if ch in "\"'":
            inq = not inq
            if not inq and buf:
                parts.append("".join(buf).strip()); buf = []
            continue
        buf.append(ch)
    if buf: parts.append("".join(buf).strip())
    terms: List[str] = []
    for p in parts:
        p = p.strip()
        if not p: continue
        if " " in p: terms.append(p)
        else: terms.extend([w for w in p.split() if w])
    seen = set(); uniq = []
    for t in terms:
        tl = t.lower()
        if tl not in seen:
            seen.add(tl); uniq.append(t)
    return uniq

def _highlight_context_html(context: str, terms: Iterable[str]) -> str:
    """
    Accent-insensitive highlight com cores distintas para cada termo:
      1) constrói versão sem acentos + mapa de índices
      2) encontra matches no string sem acentos (frases/wildcards)
      3) mapeia intervalos para índices no original
      4) injeta <span class="highlight"> com cor específica no texto ORIGINAL (com acentos)
    """
    if not context:
        return ""
    raw = context
    acc_less, idx_map = _token_spans_for_accentless_map(raw)
    if not acc_less:
        return html.escape(raw)

    # Converter termos para lista e criar paleta de cores
    terms_list = list(terms or [])
    if not terms_list:
        return html.escape(raw)
    
    # Paleta de cores para cada termo (cores distintas e acessíveis)
    color_palette = [
        "#fff59d",  # Pale yellow
        "#ffcc80",  # Orange
        "#f48fb1",  # Pink
        "#90caf9",  # Light blue
        "#a5d6a7",  # Light green
        "#ce93d8",  # Light purple
        "#ffab91",  # Light red-orange
        "#b0bec5",  # Light blue-gray
        "#ffe082",  # Amber
        "#c5e1a5",  # Light lime
        "#b39ddb",  # Light indigo
        "#ef9a9a",  # Light red
        "#81c784",  # Green
        "#64b5f6",  # Blue
        "#ffb74d",  # Deep orange
    ]
    
    # Mapear cada termo para uma cor
    term_colors = {}
    for idx, term in enumerate(terms_list):
        term_key = (term or "").lower().strip()
        if term_key:
            term_colors[term_key] = color_palette[idx % len(color_palette)]

    pats = _build_term_patterns(terms_list)
    # Matches com informação do termo (s, e, term_index)
    matches: List[Tuple[int, int, int]] = []

    for term_idx, rx in enumerate(pats):
        for m in rx.finditer(acc_less):
            s, e = m.start(), m.end()
            # mapear de volta
            s0 = idx_map[s]
            e0 = idx_map[e - 1] + 1  # fim exclusivo
            matches.append((s0, e0, term_idx))

    if not matches:
        return html.escape(raw)

    # Ordenar matches por posição
    matches.sort(key=lambda x: (x[0], x[1]))
    
    # Processar matches e aplicar cores (evitar overlaps complexos)
    out: List[str] = []
    last = 0
    for s, e, term_idx in matches:
        if s < last:
            # Overlap - usar a cor do primeiro termo encontrado
            continue
        if s > last:
            out.append(html.escape(raw[last:s]))
        
        # Obter cor do termo
        term = terms_list[term_idx] if term_idx < len(terms_list) else terms_list[0]
        term_key = (term or "").lower().strip()
        color = term_colors.get(term_key, color_palette[0])
        
        out.append(f'<span class="highlight" style="background-color: {color};">{html.escape(raw[s:e])}</span>')
        last = e
    
    if last < len(raw):
        out.append(html.escape(raw[last:]))

    return "".join(out)

# =========================
# Escrita principal
# =========================
def save_results(
    output_file: str,
    results: List[Any],
    show_duplicates: bool = False,
    output_format: str = "html",
    *,
    normalize=None,   # opcional: callable; NÃO usado para exibição (evita perder acentos)
    highlight=None,   # opcional: callable externo; se None usa o robusto local
    **kwargs          # compat: ignora extras antigos (ex.: highlight_context_html=...)
) -> None:
    """
    Guarda resultados (HTML/TXT) com layout centrado, dark-mode e highlight robusto.
    Compatível com a chamada do teu app.py (parâmetro 'highlight' e fallback legacy).
    """
    # validar formato
    if output_format not in ("txt", "html", "csv", "xlsx"):
        print(f"Warning: Invalid output format '{output_format}'. Defaulting to 'txt'.")
        output_format = "txt"

    # extensão
    if output_format == "html" and not output_file.lower().endswith(".html"):
        output_file = os.path.splitext(output_file)[0] + ".html"
    elif output_format == "csv" and not output_file.lower().endswith(".csv"):
        output_file = os.path.splitext(output_file)[0] + ".csv"
    elif output_format == "xlsx" and not output_file.lower().endswith(".xlsx"):
        output_file = os.path.splitext(output_file)[0] + ".xlsx"

    # Highlighter efetivo (suporta nome antigo)
    if highlight is None and "highlight_context_html" in kwargs:
        highlight = kwargs["highlight_context_html"]
    _hl_fn = highlight or _highlight_context_html

    # Agrupar por ficheiro e deduplicar por (context, query)
    grouped = defaultdict(list)
    seen = set()
    for r in results or []:
        key = _unique_key_for_dedup(r)
        if not show_duplicates and key in seen:
            continue
        seen.add(key)
        grouped[_as_attr(r, "filepath", "") or ""].append(r)

    # Ordenar ficheiros por max(score)
    sorted_files = sorted(
        grouped.items(),
        key=lambda kv: max((_as_attr(x, "relevance_score", 0.0) or 0.0) for x in kv[1]),
        reverse=True
    )

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # =========================
    # TXT
    # =========================
    if output_format == "txt":
        with open(output_file, "w", encoding="utf-8", errors="replace") as f:
            f.write("=" * 80 + "\n")
            f.write(f"DOCUMENT SEARCH RESULTS - Generated: {timestamp}\n")
            f.write("=" * 80 + "\n\n")

            # Summary
            unique_files = len(sorted_files)
            unique_queries = len(set(_as_attr(r, "query", "") for r in (results or [])))
            f.write("SUMMARY:\n")
            f.write(f"  Total Matches: {len(results or [])}\n")
            f.write(f"  Unique Files:  {unique_files}\n")
            f.write(f"  Search Queries:{unique_queries}\n")
            f.write("-" * 80 + "\n\n")

            for file_num, (filepath, file_results) in enumerate(sorted_files, 1):
                filename = os.path.basename(filepath)
                directory = os.path.dirname(filepath)
                abs_path = os.path.abspath(filepath)
                file_uri = f"file://{abs_path.replace(' ', '%20')}"
                f.write(f"FILE #{file_num}: {filename}\n")
                f.write(f"  Location: {directory}\n")
                f.write(f"  Full Path: {abs_path}\n")
                f.write(f"  Link: {file_uri}\n")
                f.write("-" * 80 + "\n")

                file_results.sort(key=lambda x: _as_attr(x, "relevance_score", 0.0) or 0.0, reverse=True)
                for match_num, r in enumerate(file_results, 1):
                    score = float(_as_attr(r, "relevance_score", 0.0) or 0.0)
                    location = _as_attr(r, "location", "")
                    label, page_param = _position_label(filepath, location)
                    f.write(f"\n  MATCH #{match_num} [Score: {score:.2f}]\n")
                    f.write(f"  Position: {label}\n")
                    if filepath.lower().endswith(".pdf") and page_param:
                        f.write(f"  Link: {file_uri}#page={page_param}\n")
                    else:
                        f.write(f"  Link: {file_uri}\n")
                    q = _as_attr(r, "query", "")
                    f.write(f"  Query: \"{q}\"\n\n")

                    ctx = _display_norm(_as_attr(r, "context", "") or "")
                    f.write("  " + "-" * 70 + "\n")
                    f.write("  CONTEXT:\n")
                    # wrap simples
                    col = 74
                    line = "    "
                    for w in ctx.split():
                        if len(line) + len(w) + 1 > col:
                            f.write(line + "\n")
                            line = "    " + w
                        else:
                            line += (" " if line.strip() else "") + w
                    if line.strip():
                        f.write(line + "\n")
                    f.write("  " + "-" * 70 + "\n")
                f.write("\n" + "=" * 80 + "\n\n")

            f.write("SEARCH COMPLETE\n")
            f.write("-" * 80 + "\n")
            f.write("Note: Results are sorted by relevance score (higher = better match).\n")
            f.write("      Some applications may not open file:// links directly.\n")
        return

    # =========================
    # CSV Export
    # =========================
    if output_format == "csv":
        try:
            import csv as csv_module
            with open(output_file, 'w', encoding='utf-8', newline='', errors='replace') as f:
                writer = csv_module.writer(f)
                # Header
                writer.writerow([
                    'File #', 'Filename', 'Directory', 'Full Path', 'Match #',
                    'Relevance Score', 'Position', 'Query', 'Context'
                ])
                
                # Data rows
                for file_num, (filepath, file_results) in enumerate(sorted_files, 1):
                    filename = os.path.basename(filepath)
                    directory = os.path.dirname(filepath)
                    abs_path = os.path.abspath(filepath)
                    
                    file_results.sort(key=lambda x: _as_attr(x, "relevance_score", 0.0) or 0.0, reverse=True)
                    
                    for match_num, r in enumerate(file_results, 1):
                        score = float(_as_attr(r, "relevance_score", 0.0) or 0.0)
                        location = _as_attr(r, "location", "")
                        label, _ = _position_label(filepath, location)
                        q = _as_attr(r, "query", "") or ""
                        ctx = _display_norm(_as_attr(r, "context", "") or "")
                        
                        writer.writerow([
                            file_num, filename, directory, abs_path, match_num,
                            f"{score:.4f}", label, q, ctx
                        ])
        except Exception as e:
            print(f"CSV export error: {e}")
            # Fallback to TXT
            output_format = "txt"
            output_file = os.path.splitext(output_file)[0] + ".txt"
        else:
            return

    # =========================
    # Excel Export
    # =========================
    if output_format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Search Results"
            
            # Header style
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Headers
            headers = ['File #', 'Filename', 'Directory', 'Full Path', 'Match #',
                      'Relevance Score', 'Position', 'Query', 'Context']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            row_num = 2
            for file_num, (filepath, file_results) in enumerate(sorted_files, 1):
                filename = os.path.basename(filepath)
                directory = os.path.dirname(filepath)
                abs_path = os.path.abspath(filepath)
                
                file_results.sort(key=lambda x: _as_attr(x, "relevance_score", 0.0) or 0.0, reverse=True)
                
                for match_num, r in enumerate(file_results, 1):
                    score = float(_as_attr(r, "relevance_score", 0.0) or 0.0)
                    location = _as_attr(r, "location", "")
                    label, _ = _position_label(filepath, location)
                    q = _as_attr(r, "query", "") or ""
                    ctx = _display_norm(_as_attr(r, "context", "") or "")
                    
                    ws.cell(row=row_num, column=1, value=file_num)
                    ws.cell(row=row_num, column=2, value=filename)
                    ws.cell(row=row_num, column=3, value=directory)
                    ws.cell(row=row_num, column=4, value=abs_path)
                    ws.cell(row=row_num, column=5, value=match_num)
                    ws.cell(row=row_num, column=6, value=score)
                    ws.cell(row=row_num, column=7, value=label)
                    ws.cell(row=row_num, column=8, value=q)
                    ws.cell(row=row_num, column=9, value=ctx)
                    
                    # Auto-adjust column widths
                    for col_idx in range(1, 10):
                        col_letter = get_column_letter(col_idx)
                        max_length = 0
                        for row in ws[col_letter]:
                            try:
                                if len(str(row.value)) > max_length:
                                    max_length = len(str(row.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 100)
                        ws.column_dimensions[col_letter].width = adjusted_width
                    
                    row_num += 1
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            wb.save(output_file)
        except ImportError:
            print("Excel export error: 'openpyxl' not installed. Install with: pip install openpyxl")
            # Fallback to CSV
            output_format = "csv"
            output_file = os.path.splitext(output_file)[0] + ".csv"
        except Exception as e:
            print(f"Excel export error: {e}")
            # Fallback to CSV
            output_format = "csv"
            output_file = os.path.splitext(output_file)[0] + ".csv"
        else:
            return

    # =========================
    # HTML
    # =========================
    with open(output_file, "w", encoding="utf-8", errors="replace") as f:
        # Cabeçalho HTML (sem “coluna estreita”; largura fluida até 1400px)
        f.write("""<!DOCTYPE html>
<html lang="pt">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>text-seeker — Search Results</title>
<style>
*{box-sizing:border-box}
:root{
  --bg:#ffffff; --panel:#f7f7f9; --text:#111827; --muted:#58606b;
  --border:#e5e7eb; --accent:#16a34a; --hi:#fff59d;
}
@media (prefers-color-scheme: dark){
  :root{
    --bg:#0b0c10; --panel:#111827; --text:#e5e7eb; --muted:#9aa4b2;
    --border:#2a2f3a; --accent:#22c55e; --hi:#4b5563;
  }
}
html,body{height:100%}
body{
  font-family: system-ui,-apple-system,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
  background:var(--bg); color:var(--text); line-height:1.7; margin:0;
}
.page{
  max-width: min(2500px, 200vw);
  margin:10px auto 80px;
  padding:0 clamp(16px, 3vw, 48px);
}
.header,.footer,.summary,.file{
  background:var(--panel); border:1px solid var(--border); border-radius:12px;
}
.header,.footer{ padding:20px; }
.summary{ padding:14px 18px; margin:16px 0; }
.file{ margin:24px 0; padding:16px 18px; }
.legal-footer{ margin-top:28px; padding-top:18px; border-top:1px solid var(--border); font-size:0.92em; color:var(--muted); }
.legal-footer h2{ font-size:1.05em; margin:1em 0 0.4em; color:var(--text); }
.legal-footer a{ color:var(--accent); }
.file-header{ padding:0 0 12px 0; margin:0 0 12px 0; border-bottom:1px solid var(--border); }
.match{ border-left:4px solid var(--accent); padding-left:12px; margin:14px 0; }
.context{
  background: transparent;
  padding: 0;
  border-radius: 0;
  white-space: normal;          /* em vez de pre-wrap */
  text-align: justify;
  text-justify: inter-word;
  hyphens: manual;              /* evita hífens automáticos */
  overflow-wrap: break-word;    /* só quebra palavras muito longas */
  line-height: 1.7;
}

.highlight{
  padding:0 2px; border-radius:3px; font-weight:600;
  text-decoration: underline; text-decoration-thickness: 2px; text-underline-offset: 2px;
}
/* Cor padrão (fallback) - será sobrescrita por inline styles quando especificado */
.highlight:not([style*="background-color"]) {
  background:var(--hi);
}
hr{ border:0; border-top:1px solid var(--border); margin:20px 0 }
a{ color:#0645ad; text-decoration:none } a:hover{text-decoration:underline}
.score{ color:var(--muted); font-style:italic }
.file-actions{ margin-top:10px }
.btn-open-file,.btn-copy-path{
  display:inline-block; padding:6px 10px; color:#fff; border-radius:6px; text-decoration:none; font-size:14px;
}
.btn-open-file{ background:var(--accent); margin-right:10px }
.btn-copy-path{ background:#2563eb }
.os-path{
  font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;
  background:#00000008; padding:2px 6px; border:1px solid var(--border); border-radius:6px
}
h1,h2,h3{ margin:0 0 8px }
.small{ font-size:12px; color:var(--muted) }
</style>
<script>
function copyToClipboard(text){
  navigator.clipboard.writeText(text).then(()=>{ alert('Path copied to clipboard!'); })
  .catch(err=>{ console.error('Error copying text: ', err); });
}
function openWithDefaultApp(filePath,pageNum){
  if(pageNum && filePath.toLowerCase().endsWith('.pdf')){
    const link=document.createElement('a');
    link.href='file://'+filePath+'#page='+pageNum;
    link.target='_blank';
    document.body.appendChild(link); link.click(); document.body.removeChild(link);
    return;
  }
  const ok=confirm('Open this file with the default application?\\n'+filePath);
  if(ok){
    const link=document.createElement('a');
    link.href='file://'+filePath; link.target='_blank';
    document.body.appendChild(link); link.click(); document.body.removeChild(link);
  }
}
</script>
</head>
<body>
<div class="page">
""")

        # Cabeçalho
        f.write('<div class="header">\n')
        f.write('<h1>text-seeker — Search Results</h1>\n')
        f.write(f'<p class="small">Generated: {timestamp}</p>\n')
        f.write('</div>\n')

        # Summary
        unique_files = len(sorted_files)
        unique_queries = len(set(_as_attr(r, "query", "") for r in (results or [])))
        f.write('<div class="summary">\n')
        f.write('<h2>Summary</h2>\n<ul>\n')
        f.write(f'  <li><strong>Total Matches:</strong> {len(results or [])}</li>\n')
        f.write(f'  <li><strong>Unique Files:</strong> {unique_files}</li>\n')
        f.write(f'  <li><strong>Search Queries:</strong> {unique_queries}</li>\n')
        f.write('</ul>\n</div>\n')

        # Por ficheiro
        for file_num, (filepath, file_results) in enumerate(sorted_files, 1):
            filename = os.path.basename(filepath)
            directory = os.path.dirname(filepath)
            abs_path = os.path.abspath(filepath)
            file_uri = f"file://{abs_path.replace(' ', '%20')}"
            escaped_for_js = _escape_path_for_js(abs_path)

            f.write(f'<div class="file" id="file-{file_num}">\n')
            f.write('<div class="file-header">\n')
            f.write(f'<h2>File #{file_num}: {html.escape(filename)}</h2>\n')
            f.write(f'<p><strong>Location:</strong> {html.escape(directory)}</p>\n')
            f.write(f'<p><strong>Full Path:</strong> <span class="os-path">{html.escape(abs_path)}</span></p>\n')
            f.write('<div class="file-actions">\n')
            f.write(f'  <a href="javascript:void(0)" onclick="openWithDefaultApp(\'{escaped_for_js}\', null)" class="btn-open-file">Open File</a>\n')
            f.write(f'  <a href="javascript:void(0)" onclick="copyToClipboard(\'{escaped_for_js}\')" class="btn-copy-path">Copy Path</a>\n')
            f.write('</div>\n</div>\n')

            # Ordenar matches
            file_results.sort(key=lambda x: _as_attr(x, "relevance_score", 0.0) or 0.0, reverse=True)

            for match_num, r in enumerate(file_results, 1):
                score = float(_as_attr(r, "relevance_score", 0.0) or 0.0)
                location = _as_attr(r, "location", "")
                label, page_param = _position_label(filepath, location)
                q = _as_attr(r, "query", "") or ""

                # normalizar/realçar contexto (exibição preserva acentos)
                try:
                    ctx_raw = _display_norm(_as_attr(r, "context", "") or "")
                except Exception:
                    ctx_raw = _as_attr(r, "context", "") or ""
                try:
                    terms = _collect_terms_for_highlight(q)
                    ctx_html = _hl_fn(ctx_raw, terms)
                except Exception:
                    ctx_html = html.escape(ctx_raw)

                f.write(f'<div class="match" id="match-{file_num}-{match_num}">\n')
                f.write(f'<h3>Match #{match_num} <span class="score">[Score: {score:.2f}]</span></h3>\n')
                if filepath.lower().endswith(".pdf") and page_param:
                    f.write(f'<p><strong>Position:</strong> {html.escape(label)} ')
                    f.write(f'<a href="javascript:void(0)" onclick="openWithDefaultApp(\'{escaped_for_js}\', {page_param})">(Open to Page)</a></p>\n')
                else:
                    f.write(f'<p><strong>Position:</strong> {html.escape(label)}</p>\n')
                f.write(f'<p><strong>Query:</strong> "{html.escape(q)}"</p>\n')
                f.write(f'<div class="context">{ctx_html}</div>\n')
                f.write('</div>\n')

            f.write('</div>\n')  # .file

        # Rodapé
        from legal import COPYRIGHT_HTML
        f.write("""<div class="footer">
<p>Search complete. Results are sorted by relevance score (higher = better match).</p>
<p>Highlighted terms appear underlined and shaded.</p>
<p>Click "Open File" to open the document, or "Copy Path" to copy the file path to clipboard.</p>
<p>For PDF files, you can open directly to the specific page.</p>
""")
        f.write(COPYRIGHT_HTML)
        f.write("""
</div>

</div> <!-- fecha .page -->

<script>
document.addEventListener('DOMContentLoaded', function() {
  if (navigator.userAgent.includes('Firefox')) {
    const links = document.querySelectorAll('a[href^="file://"]');
    links.forEach(link => {
      const originalHref = link.getAttribute('href');
      link.setAttribute('data-href', originalHref);
      link.setAttribute('href', 'javascript:void(0)');
      link.addEventListener('click', function() {
        alert('Firefox blocks direct file:// links. Please copy the path and open the file manually.');
        copyToClipboard(originalHref.replace('file://', ''));
      });
    });
  }
});
</script>
</body>
</html>""")


def _sanitize_folder_name(path: str) -> str:
    """Nome seguro para ficheiro a partir do path da pasta."""
    name = os.path.basename(os.path.normpath(path))
    if not name:
        name = "root"
    # Remover caracteres inválidos
    for c in '\\/:*?"<>|':
        name = name.replace(c, "_")
    return name[:64] if len(name) > 64 else name


def save_results_per_folder(
    output_dir: str,
    output_stem: str,
    results: List[Any],
    output_format: str = "html",
) -> dict:
    """
    Guarda resultados em ficheiros separados por pasta (root_folder).
    Retorna {root_folder: caminho_do_ficheiro} para construção do índice.
    """
    grouped: dict = defaultdict(list)
    for r in results or []:
        root = _as_attr(r, "root_folder", "") or ""
        grouped[root].append(r)

    saved: dict = {}
    ext = {"html": ".html", "csv": ".csv", "xlsx": ".xlsx", "txt": ".txt"}.get(output_format, ".html")
    for root_folder, folder_results in grouped.items():
        if not folder_results:
            continue
        safe_name = _sanitize_folder_name(root_folder)
        out_path = os.path.join(output_dir, f"{output_stem}_{safe_name}{ext}")
        save_results(out_path, folder_results, show_duplicates=False, output_format=output_format)
        saved[root_folder] = out_path
    return saved


def save_results_chunked(
    output_path: str,
    results: List[Any],
    max_per_file: int,
    output_format: str = "html",
) -> None:
    """
    Divide resultados em vários ficheiros (ex.: máx. 100 por ficheiro).
    Cria output_1.html, output_2.html, ... + output_INDEX.html com links.
    """
    if not results or max_per_file <= 0:
        return
    out_dir = os.path.dirname(output_path)
    out_dir = out_dir or "."
    base, ext = os.path.splitext(os.path.basename(output_path))
    saved_paths: List[str] = []
    for i in range(0, len(results), max_per_file):
        chunk = results[i : i + max_per_file]
        part_num = (i // max_per_file) + 1
        chunk_path = os.path.join(out_dir, f"{base}_{part_num}{ext}")
        save_results(chunk_path, chunk, show_duplicates=False, output_format=output_format)
        saved_paths.append(chunk_path)
    # Índice
    idx_path = os.path.join(out_dir, f"{base}_INDEX.html")
    _save_chunk_index(idx_path, saved_paths, len(results), max_per_file)


def _save_chunk_index(index_path: str, saved_paths: List[str], total: int, max_per: int) -> None:
    from datetime import datetime
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = os.path.dirname(index_path)
    lines = [
        "<!DOCTYPE html><html lang='pt'><head><meta charset='UTF-8'>",
        "<title>Índice de resultados (vários ficheiros)</title>",
        "<style>body{font-family:system-ui;max-width:700px;margin:40px auto;padding:20px}",
        ".item{margin:10px 0;padding:10px;background:#f7f7f9;border-radius:8px}",
        "a{color:#16a34a;text-decoration:none}a:hover{text-decoration:underline}</style></head><body>",
        "<h1>Índice de resultados</h1>",
        f"<p>Total: {total} resultados em {len(saved_paths)} ficheiro(s) (máx. {max_per} por ficheiro).</p>",
        f"<p class='small'>Gerado: {ts}</p><ul>",
    ]
    for idx, path in enumerate(saved_paths, 1):
        rel = os.path.relpath(path, base) if base else os.path.basename(path)
        start = (idx - 1) * max_per + 1
        end = min(idx * max_per, total)
        lines.append(f"<li class='item'><a href='{html.escape(rel)}'>Parte {idx}</a> (resultados {start}–{end})</li>")
    lines.append("</ul></body></html>")
    with open(index_path, "w", encoding="utf-8", errors="replace") as f:
        f.write("\n".join(lines))


def save_index_html(index_path: str, saved: dict, roots: List[str]) -> None:
    """Cria ficheiro índice HTML com links para cada ficheiro de resultados por pasta."""
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = os.path.dirname(index_path)
    lines = [
        "<!DOCTYPE html><html lang='pt'><head><meta charset='UTF-8'>",
        "<title>Índice de resultados</title>",
        "<style>body{font-family:system-ui;max-width:800px;margin:40px auto;padding:20px}",
        ".item{margin:12px 0;padding:12px;background:#f7f7f9;border-radius:8px}",
        "a{color:#16a34a;text-decoration:none}a:hover{text-decoration:underline}",
        "</style></head><body>",
        "<h1>Índice de resultados por pasta</h1>",
        f"<p class='small'>Gerado: {timestamp}</p>",
        "<ul>",
    ]
    for root in roots:
        if root in saved:
            rel = os.path.relpath(saved[root], base) if base else os.path.basename(saved[root])
            name = _sanitize_folder_name(root)
            lines.append(f"<li class='item'><a href='{html.escape(rel)}'><strong>{html.escape(name)}</strong></a><br><small>{html.escape(root)}</small></li>")
    from legal import COPYRIGHT_HTML
    lines.extend(["</ul>", COPYRIGHT_HTML, "</body></html>"])
    with open(index_path, "w", encoding="utf-8", errors="replace") as f:
        f.write("\n".join(lines))
