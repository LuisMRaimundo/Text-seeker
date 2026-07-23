# text_extract.py — fast full-document text extraction (indexing + BM25 corpus)
from __future__ import annotations

import os
import re
from typing import Callable, Dict, Optional, Set

from search_rdf import RDF_EXTS
from search_ebook import EBOOK_EXTS

__all__ = ["extract_document_text", "should_index_extension", "effective_extension"]

Normalizer = Callable[[str], str]

_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif", ".webp"}
_JSON_EXTS = {".json", ".jsonl"}


def effective_extension(filepath: str, ext: Optional[str] = None) -> str:
    """Return a lowercased extension, including compound ones like .fb2.zip."""
    lower = (filepath or "").lower()
    if lower.endswith(".fb2.zip"):
        return ".fb2.zip"
    if ext:
        return ext.lower()
    return os.path.splitext(lower)[1]


def should_index_extension(ext: str, file_types: Dict[str, bool]) -> bool:
    ext = (ext or "").lower()
    if file_types.get("txt") and ext in {".txt", ".log"}:
        return True
    if file_types.get("md") and ext == ".md":
        return True
    if file_types.get("csv") and ext == ".csv":
        return True
    if file_types.get("docx") and ext == ".docx":
        return True
    if file_types.get("html") and ext in {".html", ".htm"}:
        return True
    if file_types.get("excel") and ext in {".xlsx", ".xls"}:
        return True
    if file_types.get("pdf") and ext == ".pdf":
        return True
    if file_types.get("image") and ext in _IMAGE_EXTS:
        return True
    if file_types.get("json") and ext in _JSON_EXTS:
        return True
    if file_types.get("ttl") and ext in RDF_EXTS:
        return True
    if file_types.get("ebook") and ext in EBOOK_EXTS:
        return True
    return False


def _read_text_file(path: str) -> str:
    for enc in ("utf-8", "utf-8-sig", "utf-16", "latin-1", "cp1252"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except Exception:
            return ""
    return ""


def _extract_pdf_fast(path: str, max_pages: int = 0) -> str:
    parts: list[str] = []
    try:
        import fitz
        doc = fitz.open(path)
        n = doc.page_count if max_pages <= 0 else min(doc.page_count, max_pages)
        for i in range(n):
            parts.append(doc.load_page(i).get_text("text") or "")
        doc.close()
        return "\n".join(parts)
    except Exception:
        pass
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(path)
        n = len(reader.pages) if max_pages <= 0 else min(len(reader.pages), max_pages)
        for i in range(n):
            parts.append(reader.pages[i].extract_text() or "")
    except Exception:
        pass
    return "\n".join(parts)


def _extract_html_fast(path: str) -> str:
    try:
        with open(path, "rb") as f:
            raw = f.read()
    except Exception:
        return ""
    try:
        from bs4 import BeautifulSoup, Comment
    except Exception:
        return _read_text_file(path)
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            html = raw.decode(enc, errors="replace")
            break
        except Exception:
            html = raw.decode("utf-8", errors="replace")
    for parser in ("lxml", "html5lib", "html.parser"):
        try:
            soup = BeautifulSoup(html, parser)
            break
        except Exception:
            soup = None
    if soup is None:
        return ""
    root = soup.body if soup.body else soup
    for tag in root.find_all(["script", "style", "noscript", "iframe", "object"]):
        tag.decompose()
    for c in root.find_all(string=lambda t: isinstance(t, Comment)):
        c.extract()
    return root.get_text(separator="\n", strip=True)


def _extract_docx_fast(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text)
    except Exception:
        return ""


def _extract_excel_fast(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    chunks: list[str] = []
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            chunks.append(str(cell))
            wb.close()
            return "\n".join(chunks)
        except Exception:
            return ""
    try:
        import xlrd
        book = xlrd.open_workbook(path)
        for sheet in book.sheets():
            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    val = sheet.cell_value(r, c)
                    if val not in ("", None):
                        chunks.append(str(val))
        return "\n".join(chunks)
    except Exception:
        return ""


def extract_document_text(
    filepath: str,
    ext: str,
    file_types: Dict[str, bool],
    normalize: Optional[Normalizer] = None,
    *,
    ocr_image_fn=None,
    index_max_pdf_pages: int = 0,
) -> str:
    """
    Extract normalized full-document text for indexing and BM25 (fast path; PDF without OCR).
    """
    normalize = normalize or (lambda s: s)
    ext = effective_extension(filepath, ext)
    raw = ""

    if file_types.get("txt") and ext in {".txt", ".log"}:
        raw = _read_text_file(filepath)
    elif file_types.get("md") and ext == ".md":
        raw = _read_text_file(filepath)
        raw = re.sub(r"```[\s\S]*?```", " ", raw)
        raw = re.sub(r"`[^`]+`", " ", raw)
    elif file_types.get("csv") and ext == ".csv":
        raw = _read_text_file(filepath)
    elif file_types.get("docx") and ext == ".docx":
        raw = _extract_docx_fast(filepath)
    elif file_types.get("html") and ext in {".html", ".htm"}:
        raw = _extract_html_fast(filepath)
    elif file_types.get("excel") and ext in {".xlsx", ".xls"}:
        raw = _extract_excel_fast(filepath)
    elif file_types.get("pdf") and ext == ".pdf":
        raw = _extract_pdf_fast(filepath, max_pages=index_max_pdf_pages)
    elif file_types.get("image") and ext in _IMAGE_EXTS and ocr_image_fn:
        try:
            raw = ocr_image_fn(filepath) or ""
        except Exception:
            raw = ""
    elif file_types.get("json") and ext in _JSON_EXTS:
        from search_json import extract_json_text
        raw = extract_json_text(filepath)
    elif file_types.get("ttl") and ext in RDF_EXTS:
        from search_rdf import extract_rdf_text
        raw = extract_rdf_text(filepath)
    elif file_types.get("ebook") and ext in EBOOK_EXTS:
        from search_ebook import extract_ebook_text
        raw = extract_ebook_text(filepath)

    if not raw:
        return ""
    return normalize(raw)
