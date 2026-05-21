# search_excel.py — Excel file search support (XLSX, XLS)
from __future__ import annotations
from typing import Callable, List, Any, Optional
import os

Normalizer = Callable[[str], str]
Evaluator = Callable[..., List[dict]]

__all__ = ["search_in_excel_file"]

def search_in_excel_file(
    filepath: str,
    parser: Any,
    context_size: int,
    normalize: Optional[Normalizer] = None,
    evaluate_text: Optional[Evaluator] = None,
) -> List[dict]:
    """
    Lê ficheiros Excel (.xlsx, .xls) e aplica a avaliação.
    
    Extrai texto de todas as células de todas as sheets.
    Cada célula não-vazia é avaliada separadamente.
    
    Args:
        filepath: Caminho para o ficheiro Excel
        parser: BooleanSearchParser instance
        context_size: Tamanho da janela de contexto
        normalize: Função de normalização (opcional)
        evaluate_text: Função de avaliação (obrigatória)
        
    Returns:
        Lista de dicts com resultados
    """
    if evaluate_text is None:
        raise ValueError("evaluate_text callable é obrigatório.")
    normalize = normalize or (lambda s: s)
    
    results: List[dict] = []
    
    # Tentar openpyxl primeiro (para .xlsx)
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("Excel Error: dependência 'openpyxl' em falta. Instale com: pip install openpyxl")
            return []
        
        try:
            wb = load_workbook(filepath, data_only=True, read_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Iterar por todas as células
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for col_idx, cell_value in enumerate(row, start=1):
                        if cell_value is None:
                            continue
                        
                        # Converter para string
                        cell_text = str(cell_value).strip()
                        if not cell_text:
                            continue
                        
                        try:
                            text = normalize(cell_text)
                            if text:
                                # Location: sheet_name, row, col
                                location_str = f"{sheet_name}:{row_idx}:{col_idx}"
                                hits = evaluate_text(
                                    text, filepath, parser, context_size,
                                    location=location_str,
                                    location_type="cell"
                                )
                                if hits:
                                    results.extend(hits)
                        except Exception as e:
                            print(f"Excel evaluate error ({filepath}, {sheet_name}, R{row_idx}C{col_idx}): {e}")
                            continue
            
            wb.close()
            
        except Exception as e:
            print(f"Excel Error {filepath}: {e}")
            return []
    
    elif ext == '.xls':
        # Para .xls antigos, usar xlrd
        try:
            import xlrd
        except ImportError:
            print("Excel Error: dependência 'xlrd' em falta. Instale com: pip install xlrd")
            return []
        
        try:
            wb = xlrd.open_workbook(filepath)
            
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                sheet_name = sheet.name
                
                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        cell_value = cell.value
                        
                        if cell_value is None:
                            continue
                        
                        # Converter para string
                        cell_text = str(cell_value).strip()
                        if not cell_text:
                            continue
                        
                        try:
                            text = normalize(cell_text)
                            if text:
                                # Location: sheet_name, row, col
                                location_str = f"{sheet_name}:{row_idx+1}:{col_idx+1}"
                                hits = evaluate_text(
                                    text, filepath, parser, context_size,
                                    location=location_str,
                                    location_type="cell"
                                )
                                if hits:
                                    results.extend(hits)
                        except Exception as e:
                            print(f"Excel evaluate error ({filepath}, {sheet_name}, R{row_idx+1}C{col_idx+1}): {e}")
                            continue
            
        except Exception as e:
            print(f"Excel Error {filepath}: {e}")
            return []
    
    return results
